from __future__ import annotations

from collections import Counter
from collections.abc import Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, TypeVar

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.models import (
    ActionItem,
    AIExplanation,
    DecisionRecord,
    ImportBatch,
    ImportIssue,
    ProcessingJob,
    RuleVersion,
    SpuSnapshot,
    new_id,
)
from app.rules import RULE_VERSION, RuleInput, decide

ALIASES: dict[str, tuple[str, ...]] = {
    "spu_id": ("SPU ID", "SPU", "链接", "链接ID", "商品链接", "商品链接ID"),
    "spu_name": ("链接名称", "商品名称", "SPU名称"),
    "store": ("店铺", "店铺名称"),
    "platform": ("平台", "渠道平台"),
    "operator_ref": ("运营", "责任运营", "运营负责人"),
    "launch_date": ("上架时间", "上架日期"),
    "net_sales": ("销售收入", "净销售额", "上月净销售额"),
    "profit_rate": ("经营准利润率", "上一周经营准利润率", "上月经营准利润率"),
    "promotion_expense": ("推广费用", "广告费用"),
    "return_count_7d": ("最近7天品退件数", "7天品退件数", "退货数量"),
    "sold_count_7d": ("最近7天已销售件数", "7天已销售件数", "销量"),
    "return_period_verified": ("品退周期已验证", "最近7天品退已验证"),
    "warehouse_qty": ("仓内库存", "仓内库存数量"),
    "in_transit_qty": ("在途库存", "在途库存数量"),
    "sales_units_14d": ("最近14天销量", "14天销量"),
}
REQUIRED_IDENTITY = ("spu_id", "spu_name", "store", "platform", "operator_ref")
ParsedValue = TypeVar("ParsedValue")


def clean_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().replace("\n", "").split())


def find_header_row(rows: Sequence[Sequence[object]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:10]):
        normalized = [clean_header(value) for value in row]
        mapping: dict[str, int] = {}
        for field, aliases in ALIASES.items():
            alias_values = {clean_header(alias) for alias in aliases}
            matches = [column for column, value in enumerate(normalized) if value in alias_values]
            if len(matches) == 1:
                mapping[field] = matches[0]
        if all(field in mapping for field in REQUIRED_IDENTITY):
            return index, mapping
    raise ValueError("未找到可唯一映射的 SPU 身份表头")


def parse_decimal(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        raise ValueError("布尔值不能作为数值")
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        result = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("数值不可解析") from exc
    return result / Decimal("100") if percent else result


def parse_int(value: object) -> int | None:
    number = parse_decimal(value)
    if number is None:
        return None
    if number != number.to_integral_value():
        raise ValueError("件数必须为整数")
    return int(number)


def parse_date(value: object) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = from_excel(value)
        return parsed.date() if isinstance(parsed, datetime) else parsed
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("日期不可解析")


def parse_bool(value: object) -> bool:
    return clean_header(value).lower() in {"是", "已验证", "true", "1", "yes"}


def json_safe(value: object) -> object:
    if isinstance(value, (datetime, date, Decimal)):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def issue(
    batch_id: str,
    row_number: int,
    field: str,
    value: object,
    severity: str,
    code: str,
    message: str,
    continues: bool,
) -> ImportIssue:
    return ImportIssue(
        batch_id=batch_id,
        source_row=row_number,
        field=field,
        original_value="" if value is None else str(value),
        severity=severity,
        code=code,
        message=message,
        continues_processing=continues,
    )


def structured_four_elements(
    snapshot: SpuSnapshot, result_action: str, inventory: str
) -> dict[str, object]:
    action_labels = {
        "clearance": "清仓",
        "stop_loss": "止损",
        "observe": "持续观察",
        "invest": "加投",
        "maintain": "维持",
        "undetermined": "无法判定",
    }
    replenishment_labels = {
        "forbid": "禁止补货",
        "replenish": "补货",
        "no_replenishment": "不补货",
        "not_generated": "不生成补货判断",
    }
    action_text = action_labels[result_action]
    replenishment_text = replenishment_labels[inventory]
    return {
        "object": {"spu_id": snapshot.spu_id, "spu_name": snapshot.spu_name},
        "problem": f"固定规则判定经营主动作为“{action_text}”。",
        "evidence": {
            "net_sales": str(snapshot.net_sales) if snapshot.net_sales is not None else None,
            "profit_rate": str(snapshot.profit_rate) if snapshot.profit_rate is not None else None,
            "return_rate_7d": (
                str(snapshot.return_rate_7d) if snapshot.return_rate_7d is not None else None
            ),
            "inventory_days": (
                str(snapshot.inventory_days) if snapshot.inventory_days is not None else None
            ),
            "quality_flags": snapshot.quality_flags,
        },
        "action": f"{action_text}；库存动作：{replenishment_text}。",
    }


def process_import(db: Session, batch_id: str) -> None:
    batch = db.get(ImportBatch, batch_id)
    if batch is None or batch.status == "ready":
        return
    if not batch.source_path:
        raise ValueError("批次源文件暂存路径不可用")
    source_path = Path(batch.source_path)
    batch.status = "validating"
    db.commit()
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        header_index, mapping = find_header_row(rows)
        data_rows = [
            row for row in rows[header_index + 1 :] if any(value is not None for value in row)
        ]
        raw_spu_ids = [
            clean_header(row[mapping["spu_id"]])
            for row in data_rows
            if mapping["spu_id"] < len(row)
        ]
        duplicates = {
            spu_id for spu_id, count in Counter(raw_spu_ids).items() if spu_id and count > 1
        }

        db.execute(delete(ImportIssue).where(ImportIssue.batch_id == batch_id))
        db.flush()
        snapshots: list[SpuSnapshot] = []
        rejected = 0
        degraded = 0
        warnings = 0
        for offset, row in enumerate(data_rows, start=header_index + 2):
            values = {
                field: row[column] if column < len(row) else None
                for field, column in mapping.items()
            }
            identity_errors = [
                field for field in REQUIRED_IDENTITY if clean_header(values.get(field)) == ""
            ]
            spu_id = clean_header(values.get("spu_id"))
            if spu_id in duplicates:
                identity_errors.append("spu_id")
            if identity_errors:
                rejected += 1
                for field in sorted(set(identity_errors)):
                    message = (
                        "批次内 SPU ID 重复"
                        if field == "spu_id" and spu_id in duplicates
                        else "必要身份字段未提供"
                    )
                    db.add(
                        issue(
                            batch_id,
                            offset,
                            field,
                            values.get(field),
                            "rejected",
                            "IDENTITY_INVALID",
                            message,
                            False,
                        )
                    )
                continue

            quality_flags: list[str] = []

            def parse_field(
                field: str,
                parser: Callable[[object], ParsedValue | None],
                row_values: Mapping[str, object],
                row_flags: list[str],
                row_offset: int,
            ) -> ParsedValue | None:
                nonlocal degraded
                raw = row_values.get(field)
                try:
                    parsed = parser(raw)
                except ValueError as exc:
                    degraded += 1
                    row_flags.append(f"{field}:invalid")
                    db.add(
                        issue(
                            batch_id,
                            row_offset,
                            field,
                            raw,
                            "degraded",
                            "FIELD_INVALID",
                            str(exc),
                            True,
                        )
                    )
                    return None
                if parsed is None:
                    degraded += 1
                    row_flags.append(f"{field}:unavailable")
                    db.add(
                        issue(
                            batch_id,
                            row_offset,
                            field,
                            raw,
                            "degraded",
                            "FIELD_UNAVAILABLE",
                            "字段未提供，相关判断停止",
                            True,
                        )
                    )
                return parsed

            launch_date = parse_field("launch_date", parse_date, values, quality_flags, offset)
            net_sales = parse_field("net_sales", parse_decimal, values, quality_flags, offset)
            profit_rate = parse_field("profit_rate", parse_decimal, values, quality_flags, offset)
            promotion_expense = parse_field(
                "promotion_expense", parse_decimal, values, quality_flags, offset
            )
            return_count = parse_field("return_count_7d", parse_int, values, quality_flags, offset)
            sold_count = parse_field("sold_count_7d", parse_int, values, quality_flags, offset)
            return_verified = parse_bool(values.get("return_period_verified"))
            return_rate: Decimal | None = None
            if return_count is not None and sold_count is not None:
                if return_count < 0 or sold_count < 0:
                    degraded += 1
                    quality_flags.append("return_rate_7d:invalid")
                    db.add(
                        issue(
                            batch_id,
                            offset,
                            "return_rate_7d",
                            "",
                            "degraded",
                            "RETURN_NEGATIVE",
                            "品退件数或销量不能为负",
                            True,
                        )
                    )
                elif sold_count == 0:
                    quality_flags.append("return_rate_7d:no_sales")
                    db.add(
                        issue(
                            batch_id,
                            offset,
                            "return_rate_7d",
                            "0",
                            "degraded",
                            "RETURN_ZERO_DENOMINATOR",
                            "无可计算销量",
                            True,
                        )
                    )
                    degraded += 1
                elif return_verified:
                    return_rate = Decimal(return_count) / Decimal(sold_count)
                else:
                    quality_flags.append("return_rate_7d:unverified_period")
                    db.add(
                        issue(
                            batch_id,
                            offset,
                            "return_rate_7d",
                            "",
                            "warning",
                            "RETURN_PERIOD_UNVERIFIED",
                            "品退数据未校验为最近 7 天",
                            True,
                        )
                    )
                    warnings += 1
            warehouse_qty = parse_field(
                "warehouse_qty", parse_decimal, values, quality_flags, offset
            )
            in_transit_qty = parse_field(
                "in_transit_qty", parse_decimal, values, quality_flags, offset
            )
            sales_units_14d = parse_field(
                "sales_units_14d", parse_decimal, values, quality_flags, offset
            )

            snapshot = SpuSnapshot(
                batch_id=batch_id,
                spu_id=spu_id,
                spu_name=clean_header(values["spu_name"]),
                store=clean_header(values["store"]),
                platform=clean_header(values["platform"]),
                operator_ref=clean_header(values["operator_ref"]),
                launch_date=launch_date,
                net_sales=net_sales,
                profit_rate=profit_rate,
                promotion_expense=promotion_expense,
                return_count_7d=return_count,
                sold_count_7d=sold_count,
                return_rate_7d=return_rate,
                return_period_verified=return_verified,
                warehouse_qty=warehouse_qty,
                in_transit_qty=in_transit_qty,
                sales_units_14d=sales_units_14d,
                raw_row=offset,
                raw_values={field: json_safe(value) for field, value in values.items()},
                quality_flags=quality_flags,
            )
            db.add(snapshot)
            snapshots.append(snapshot)

        db.flush()
        rule_version = db.get(RuleVersion, RULE_VERSION)
        if rule_version is None:
            rule_version = RuleVersion(
                version=RULE_VERSION,
                effective_from=date(2026, 1, 1),
                rules={
                    "source": "PRD详细版.md §3",
                    "priority": ["clearance", "stop_loss", "observe", "invest", "maintain"],
                },
            )
            db.add(rule_version)
            db.flush()

        batch.status = "rules_processing"
        for snapshot in snapshots:
            rule_input = RuleInput(
                business_date=batch.business_date,
                launch_date=snapshot.launch_date,
                net_sales=snapshot.net_sales,
                profit_rate=snapshot.profit_rate,
                return_rate_7d=snapshot.return_rate_7d,
                return_period_verified=snapshot.return_period_verified,
                warehouse_qty=snapshot.warehouse_qty,
                in_transit_qty=snapshot.in_transit_qty,
                sales_units_14d=snapshot.sales_units_14d,
            )
            result = decide(rule_input)
            snapshot.inventory_days = result.inventory_days
            decision_id = new_id("DEC")
            decision = DecisionRecord(
                decision_id=decision_id,
                batch_id=batch_id,
                snapshot_id=snapshot.id,
                spu_id=snapshot.spu_id,
                category=result.category,
                main_action=result.main_action,
                replenishment_action=result.replenishment_action,
                rule_version=RULE_VERSION,
                triggered_rules=list(result.triggered_rules),
                key_inputs={
                    "business_date": str(batch.business_date),
                    "period": [str(batch.period_start), str(batch.period_end)],
                    "net_sales": str(snapshot.net_sales)
                    if snapshot.net_sales is not None
                    else None,
                    "profit_rate": str(snapshot.profit_rate)
                    if snapshot.profit_rate is not None
                    else None,
                    "return_rate_7d": str(snapshot.return_rate_7d)
                    if snapshot.return_rate_7d is not None
                    else None,
                    "inventory_days": str(result.inventory_days)
                    if result.inventory_days is not None
                    else None,
                },
                four_elements={},
            )
            decision.four_elements = structured_four_elements(
                snapshot, result.main_action, result.replenishment_action
            )
            db.add(decision)
            db.flush()
            if result.main_action not in {"maintain", "undetermined"}:
                db.add(
                    ActionItem(
                        action_id=new_id("ACT"),
                        decision_id=decision_id,
                        slot="operation",
                        action_value=result.main_action,
                        owner_role="operator",
                    )
                )
            if result.replenishment_action in {"replenish", "forbid"}:
                db.add(
                    ActionItem(
                        action_id=new_id("ACT"),
                        decision_id=decision_id,
                        slot="procurement",
                        action_value=result.replenishment_action,
                        owner_role="procurement",
                    )
                )
            db.add(AIExplanation(decision_id=decision_id, status="pending"))
            db.add(
                ProcessingJob(
                    job_id=new_id("JOB"),
                    idempotency_key=f"ai:{decision_id}",
                    job_type="ai_explanation",
                    payload={"decision_id": decision_id},
                )
            )

        batch.valid_row_count = len(snapshots)
        batch.rejected_row_count = rejected
        batch.degraded_field_count = degraded
        batch.warning_count = warnings
        batch.status = "ready" if snapshots else "failed"
        batch.error_message = None if snapshots else "没有可处理的 SPU 身份行"
        batch.source_path = None
        db.commit()
    except Exception as exc:
        db.rollback()
        failed_batch = db.get(ImportBatch, batch_id)
        if failed_batch is not None:
            failed_batch.status = "failed"
            failed_batch.error_message = str(exc)[:500]
            db.commit()
        raise
    finally:
        source_path.unlink(missing_ok=True)
